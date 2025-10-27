from typing import Callable, List
import datetime

import xlrd
from openpyxl import load_workbook

from .models import AccountTransaction, CreditCardTransaction


class AccountTransactionsInput:
    """Read an .xlsx file containing bank account transactions into a list of AccountTransaction objects"""

    def __init__(self, excel_file: str, resolve_payee: Callable[[str], str], min_row: int = 11, max_col: int = 8):
        self.excel_file = excel_file
        self.resolve_payee = resolve_payee
        self.min_row = min_row
        self.max_col = max_col

    def read(self) -> List[AccountTransaction]:
        """Read an .xlsx file containing bank account transactions into a list of AccountTransaction objects"""
        workbook = load_workbook(filename=self.excel_file)
        ws = workbook.active

        transactions = []
        for row in ws.iter_rows(min_row=self.min_row, max_col=self.max_col):
            # Fields:
            # ------------------------------------------------------------
            # 0: Data_Operazione
            # 1: Data_Valuta (not used)
            # 2: Entrate
            # 3: Uscite
            # 4: Descrizione
            # 5: Descrizione_Completa
            # 6: Stato
            # 7: Moneymap
            # ------------------------------------------------------------
            date_value = row[0].value

            # Skip rows with no Data_Operazione (empty or '-'):
            # This means the transaction is not yet booked, so
            # we skip it for now. It will be booked later.
            if not date_value or date_value == '-':
                continue

            # Ensure date is a datetime object
            if isinstance(date_value, str):
                date_value = datetime.datetime.strptime(date_value, "%Y-%m-%d")
            elif not isinstance(date_value, datetime.datetime):
                # Handle date objects by converting to datetime
                if hasattr(date_value, 'year'):
                    date_value = datetime.datetime.combine(date_value, datetime.time())

            t = AccountTransaction(
                date=date_value,
                amount=row[2].value or row[3].value,
                description=row[4].value,
                description_full=row[5].value,
                state=row[6].value,
                moneymap_category=row[7].value,
                # Join description and description_full to create a
                # unique key for the payee resolver
                payee=self.resolve_payee(f"{row[4].value or ''}: {row[5].value or ''}"),
            )

            transactions.append(t)

        return transactions

class CreditCardTransactionsInput:
    """Read a credit card statement and output a list of CreditCardTransaction objects"""

    def __init__(self, excel_file: str, resolve_payee: Callable[[str], str], circuit: str = None):
        self.excel_file = excel_file
        self.resolve_payee = resolve_payee
        self.circuit = circuit

    def read(self) -> List[CreditCardTransaction]:
        """Read a credit card statement and output a list of CreditCardTransaction objects"""
        # Open the workbook
        workbook = xlrd.open_workbook(self.excel_file)

        # Select the first sheet (index 0) from the workbook
        sheet = workbook.sheet_by_index(0)

        transactions = []

        # Read data from cells with data
        for row in range(3, sheet.nrows):

            cell_value = sheet.cell_value(row, 1)
            if cell_value == "":
                continue
                
            owner = sheet.cell_value(row, 1)
            card_number = sheet.cell_value(row, 2)

            transaction_date = datetime.datetime(*xlrd.xldate_as_tuple(sheet.cell_value(row, 3), workbook.datemode))
            registration_date = datetime.datetime(*xlrd.xldate_as_tuple(sheet.cell_value(row, 4), workbook.datemode))

            description = sheet.cell_value(row, 5)
            operation_state = sheet.cell_value(row, 6)
            operation_type = sheet.cell_value(row, 7)
            circuit = sheet.cell_value(row, 8)
            transaction_type = sheet.cell_value(row, 9)
            amount = sheet.cell_value(row, 10)
            payee = self.resolve_payee(description)

            if self.circuit != "ALL" and circuit != self.circuit:
                continue

            transaction = CreditCardTransaction(
                owner=owner,
                card_number=card_number,
                transaction_date=transaction_date,
                registration_date=registration_date,
                description=description,
                operation_state=operation_state,
                operation_type=operation_type,
                circuit=circuit,
                transaction_type=transaction_type,
                amount=amount,
                payee=payee,
            )

            transactions.append(transaction)
        
        return transactions

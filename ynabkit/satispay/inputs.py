from typing import List, Callable
import datetime

from openpyxl import load_workbook

from .models import Transaction


class TransactionsInput:

    def __init__(self, excel_file: str, payee_resolver: Callable[[str], str], exclude_kinds: List[str] = None):
        self.excel_file = excel_file
        self.exclude_kinds = exclude_kinds
        self.payee_resolver = payee_resolver

    def read(self) -> List[Transaction]:
        """Read a credit card statement and output a list of Transaction objects"""
        # Open the workbook
        workbook = load_workbook(filename=self.excel_file)
        ws = workbook.active

        transactions = []
        for row in ws.iter_rows(min_row=2, max_col=8):
            # Skip if the kind is in the exclude list
            if self.exclude_kinds and row[4].value in self.exclude_kinds:
                continue

            # Fields:
            # ------------------------------------------------------------
            # 0: Date
            # 1: Name
            # 2: Description
            # 3: Amount
            # 4: Type
            # 5: Status
            # 6: Balance
            # 7: Balance after transaction
            # 8: ID (not available)
            # ------------------------------------------------------------
            
            t = Transaction(
                name=row[1].value,
                state=row[5].value,
                kind=row[4].value,
                date=row[0].value,
                amount=row[3].value,
                payee=self.payee_resolver(row[1].value),
            )

            transactions.append(t)
        return transactions

    def _parse_date(self, date_str: str) -> datetime.date:
        # try one of the following date formats
        date_formats = [
            '%d %b %Y at %I:%M:%S %p',
            '%d %b %Y at %H:%M:%S',
        ]
        for date_format in date_formats:
            try:
                return datetime.datetime.strptime(date_str, date_format)
            except ValueError:
                pass

        raise ValueError(f"Could not parse date {date_str}")

    def _parse_amount(self, amount_str: str) -> float:
        """
        Parse string amounts like "+40" or "-12,4" to float
        
        Args:
            amount_str: String representation of amount (e.g., "+40", "-12,4")
            
        Returns:
            float: The parsed amount
            
        Examples:
            >>> parse_amount("+40")
            40.0
            >>> parse_amount("-12,4")
            -12.4
            >>> parse_amount("1.234,56")
            1234.56        
        """
        return float(amount_str.replace(".", "").replace(",", "."))

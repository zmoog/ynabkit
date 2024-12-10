import datetime

import click
from openpyxl import load_workbook

from .fineco.models import AccountTransaction
from .fineco.inputs import CreditCardTransactionsInput
from .fineco.outputs import CreditCardTransactionsOutput, AccountTransactionsOutput
from .satispay.inputs import TransactionsInput
from .satispay.outputs import TransactionsOutput
from .n26.inputs import TransactionsInput as N26TransactionsInput
from .n26.outputs import TransactionsOutput as N26TransactionsOutput

@click.group()
@click.version_option()
def cli():
    "CLI tool to support data import and export from YNAB"


@cli.group()
def fineco():
    "Fineco related commands"


@cli.group()
def satispay():
    "Satispay related commands"


@cli.group()
def n26():
    "N26 related commands"


@fineco.command(name="describe-account-transactions")
@click.argument(
    "excel-file-name",
)
@click.option(
    "-o",
    "--output-format",
    help="Output format",
    type=click.Choice(["table", "csv", "json"]),
    default="table",
)
def describe_account_transactions(excel_file_name: str, output_format: str):
    "Read an .xlsx file containing bank account transactions and output the in a table or a CSV file"
    workbook = load_workbook(filename=excel_file_name)
    ws = workbook.active

    transactions = []
    for row in ws.iter_rows(min_row=8, max_col=7):
        t = AccountTransaction(
            date=datetime.datetime.strptime(row[0].value, '%d/%m/%Y').date(),
            amount=row[1].value or row[2].value,
            description=row[3].value,
            description_full=row[4].value,
            state=row[5].value,
            moneymap_category=row[6].value,
        )

        transactions.append(t)

    output = AccountTransactionsOutput(transactions, resolve_payee=resolve_payee)
    if output_format == "table":
        click.echo(output.table())
    elif output_format == "csv":
        click.echo(output.csv())
    elif output_format == "json":
        click.echo(output.json())




@fineco.command(name="describe-card-transactions")
@click.argument(
    "excel-file-name",
)
@click.option(
    "-o",
    "--output-format",
    help="Output format",
    type=click.Choice(["table", "csv", "json"]),
    default="table",
)
@click.option(
    "-c",
    "--circuit",
    help="Credit card circuit",
    type=click.Choice(["ALL", "BANCOMAT", "VISA", "MASTERCARD"]),
    default="ALL",
)
def describe_card_transactions(excel_file_name: str, output_format: str, circuit: str = None):
    "Read an .xlsx file containing credit card transactions and output the in a table or a CSV file"
 
    input = CreditCardTransactionsInput(excel_file_name, circuit=circuit)

    transactions = input.read()

    output = CreditCardTransactionsOutput(transactions, resolve_payee=resolve_payee)
    if output_format == "table":
        click.echo(output.table())
    elif output_format == "csv":
        click.echo(output.csv())
    elif output_format == "json":
        click.echo(output.json())


@satispay.command(name="describe-transactions")
@click.argument(
    "excel-file-name",
)
@click.option(
    "-o",
    "--output-format",
    help="Output format",
    type=click.Choice(["table", "csv", "json"]),
    default="table",
)
def describe_transactions(excel_file_name: str, output_format: str):
    "Read an .xlsx file containing credit card transactions and output the in a table or a CSV file"
 
    input = TransactionsInput(excel_file_name)

    transactions = input.read()

    output = TransactionsOutput(transactions, resolve_payee=resolve_payee)
    if output_format == "table":
        click.echo(output.table())
    elif output_format == "csv":
        click.echo(output.csv())
    elif output_format == "json":
        click.echo(output.json())


@n26.command(name="describe-transactions")
@click.argument(
    "csv-file-name",
)
@click.option(
    "-s",
    "--skip-header",
    help="Skip the header row",
    default=True,
)
@click.option(
    "-o",
    "--output-format",
    help="Output format",
    type=click.Choice(["table", "csv", "json"]),
    default="table",
)
def describe_n26_transactions(csv_file_name: str, skip_header: bool, output_format: str):
    "Read an .csv file containing N26 transactions and output the in a table or a CSV file"

    transactions = N26TransactionsInput(csv_file_name, skip_header=skip_header).read() 

    output = N26TransactionsOutput(transactions, payee_resolver=resolve_payee)
    if output_format == "table":
        click.echo(output.table())
    elif output_format == "csv":
        click.echo(output.csv())
    elif output_format == "json":
        click.echo(output.json())
def resolve_payee(memo: str) -> str:
    mappings = {
        "audible.it": "Audible",
        "Borello": "Borello",
        "CARREFOUR": "Carrefour",
        "Conad": "Conad",
        "Coop": "Coop",
        "Cinema Ideal": "Ideal Citiplex Torino",
        "PAYPAL *DAZN": "DAZN",
        "ELASTIC ITALY S.R.L.": "Elastic",
        "Fatna": "Fatna",
        "GOOGLE YOUTUBE": "YouTube",
        "GRAMMARLY": "Grammarly",
        "ILIAD Italia": "Iliad",
        "Focaccia Siciliana": "Na' Focaccia Siciliana",
        "LOCANDA 10038": "Pizzeria 10038",
        "Lovet": "Lovet",
        "MAXIMUM FUN MEDIA": "Maximum Fun",
        "Netflix": "Netflix",
        "Pane Amore e": "Cossa",
        "PANIFICIO FAM FAVRO": "Favro",
        "PLAYSTATION": "Sony",
        "Prime Video": "Prime Video",
        "QUALITY GROUP SOCIETA CONSORTILE": "Quality Group",
        "STEAM GAMES ": "Steam",
        "SPOTIFY": "Spotify",
        "TELECOMITALIA SPA": "TIM",
        "UnipolTech": "Unipol Tech",
        "Vodafone": "Vodafone",
        "WINDTRE RHO": "Wind",
    }

    for pattern, payee in mappings.items():
        if pattern.casefold() in memo.casefold():
            return payee
        
    return "" # default to empty string if no match is found
